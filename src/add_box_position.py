import argparse
import glob
import ntpath
import os
from functools import reduce

import pandas as pd
from fitz import fitz, Rect
from openpyxl import load_workbook, Workbook

from src.workbook import write_style_frame


def main(input_document: str, input_workbook: str, output_workbook: str):
    workbook: Workbook = load_workbook(input_workbook)  # noqa
    reading_sheet = workbook.active

    def merge_rect(a: Rect, b: Rect):
        a.includeRect(b)
        return a

    with fitz.Document(input_document) as doc:
        df = pd.DataFrame(columns=["Text", "page", "x1", "y1", "x2", "y2", "Entity", "Label"])
        for idx, row in enumerate(reading_sheet.iter_rows(min_row=2, values_only=True), start=1):
            for page in doc:
                rect_list = page.searchFor(row[0])  # noqa
                if not rect_list:
                    continue
                rect = reduce(merge_rect, rect_list)
                print(idx, row[0], page.number, rect)
                df = df.append(
                    {
                        "Text": row[0],
                        "page": page.number,
                        "x1": rect.x0,
                        "y1": rect.y0,
                        "x2": rect.x1,
                        "y2": rect.y1,
                        "Entity": row[1] if row[1] else "",
                        "Label": row[2] if row[2] else "",
                    },
                    ignore_index=True,
                )
    write_style_frame(df, output_workbook)


if __name__ == "__main__":
    parser = argparse.ArgumentParser()
    parser.add_argument("src", help="PDF source")
    parser.add_argument("dest", help="Workbook destination")
    args = parser.parse_args()

    doc_files = glob.glob(os.path.join(args.src, "*.pdf")) if os.path.isdir(args.src) else [args.src]
    workbook_files = [
        os.path.join(ntpath.dirname(doc), ntpath.basename(doc).replace("pdf", "xlsx")) for doc in doc_files
    ]
    output_workbooks = [os.path.join(args.dest, ntpath.basename(doc).replace("pdf", "xlsx")) for doc in doc_files]
    for doc, workbook, dest_file in zip(doc_files, workbook_files, output_workbooks):
        main(doc, workbook, dest_file)
